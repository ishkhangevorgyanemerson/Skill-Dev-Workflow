import argparse, json
from openpyxl import load_workbook
from mapping_rules import detect_style_from_sheet


def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def find_texts(ws, rows=(1, 2, 3, 7), max_col=25):
    vals = []
    for r in rows:
        for c in range(1, min(ws.max_column, max_col) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                vals.append(v.strip())
    return vals


def has_mapping_column(ws):
    targets = {
        "test step",
        "teststep",
        "ni evalution",
        "ni evaluation",
        "predicted test step",
    }
    for row in (1, 2, 3, 7):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row, c).value
            if isinstance(v, str) and v.strip().lower() in targets:
                return True, c, row
    return False, None, None


def candidate_score(ws, wb_names):
    style = detect_style_from_sheet(ws, wb_names, ws.title)
    header = " | ".join(find_texts(ws)).lower()
    score = 0
    reasons = []
    style_weight = {
        "LIMIT_RF": 97,
        "SPEC_UPDATE": 96,
        "E6": 96,
        "E5": 95,
        "E4": 88,
        "E3": 84,
        "E2": 90,
        "E1": 78,
        "UNKNOWN": 25,
    }
    score += style_weight.get(style, 25)
    reasons.append(f"style={style}")
    key_terms = [
        "test number",
        "test name",
        "testnamemod",
        "spec_name",
        "rf_in",
        "rf_out",
        "mipi",
        "set_freq_1",
        "test step",
        "test #",
        "ni evalution",
        "ni evaluation",
        "testtime_1site",
        "testtime_4site",
        "test id",
        "test description",
        "test item description",
        "mode",
        "waveform",
        "pin (dbm)",
        "pout (dbm)",
        "freq (mhz)",
        "power (dbm)",
    ]
    hits = sum(1 for k in key_terms if k in header)
    score += hits * 2
    reasons.append(f"header_hits={hits}")
    checked = min(ws.max_row, 120)
    num = sum(1 for r in range(1, checked + 1) if is_num(ws.cell(r, 1).value))
    ratio = num / checked if checked else 0
    score += int(ratio * 20)
    reasons.append(f"numeric_row_ratio={ratio:.2f}")
    hm, mc, mr = has_mapping_column(ws)
    return {
        "sheet": ws.title,
        "detected_style": style,
        "score": score,
        "has_mapping_column": hm,
        "mapping_column_index": mc,
        "mapping_header_row": mr,
        "suggested_action": "analyze directly"
        if hm
        else "create mapping column then analyze",
        "reasons": reasons,
    }


ap = argparse.ArgumentParser()
ap.add_argument("--input", required=True)
args = ap.parse_args()
wb = load_workbook(args.input)
rows = sorted(
    [candidate_score(wb[s], wb.sheetnames) for s in wb.sheetnames],
    key=lambda x: x["score"],
    reverse=True,
)
print(
    json.dumps({"input": args.input, "candidates": rows}, ensure_ascii=False, indent=2)
)
