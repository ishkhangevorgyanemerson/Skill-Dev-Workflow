import argparse
import json
from openpyxl import load_workbook
from mapping_rules import (
    detect_style_from_sheet,
    find_semantic_columns,
    infer_limit_rf_mapping,
    infer_spec_update_step,
    norm,
)

ap = argparse.ArgumentParser()
ap.add_argument("--input", required=True)
ap.add_argument("--sheet")
args = ap.parse_args()
wb = load_workbook(args.input, data_only=False)
ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
style = detect_style_from_sheet(ws, wb.sheetnames, ws.title)
records = []

if style == "LIMIT_RF":
    header_row = None
    for r in range(1, 10):
        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, 18)]
        if "test id" in vals and "Item" in vals and "Test Description" in vals:
            header_row = r
            break
    start = (header_row or 2) + 2
    current_section = ""
    prev = None

    def is_num(v):
        return isinstance(v, (int, float)) and not isinstance(v, bool)

    for r in range(start, ws.max_row + 1):
        tid = ws.cell(r, 1).value
        item = ws.cell(r, 2).value
        desc = ws.cell(r, 3).value
        if not is_num(tid):
            section = " ".join(
                [str(ws.cell(r, c).value or "").strip() for c in range(1, 4)]
            ).strip()
            if section:
                current_section = section
                prev = None
            continue
        band = ws.cell(r, 8).value
        tx = ws.cell(r, 9).value
        freq = ws.cell(r, 10).value
        wave = ws.cell(r, 11).value
        pwr = ws.cell(r, 12).value
        m = infer_limit_rf_mapping(
            item, desc, band, tx, freq, wave, pwr, current_section, prev
        )
        if m:
            prev = m
        records.append(
            {
                "row": r,
                "test_id": tid,
                "item": item,
                "desc": desc,
                "suggested_mapping": m,
            }
        )

elif style == "SPEC_UPDATE":
    sem = find_semantic_columns(ws)
    if sem is not None:
        cols = sem["columns"]
        hr = sem["header_row"]
        c_desc = cols.get("description")
        c_cond = cols.get("condition")
        c_mode = cols.get("mode")
        c_wave = cols.get("waveform")
        c_freq = cols.get("freq")
        c_pin = cols.get("pin")
        c_pout = cols.get("pout")
        c_test_no = cols.get("test_no", 1)
        last_mode = ""
        last_wave = ""

        for r in range(hr + 1, ws.max_row + 1):
            desc = ws.cell(r, c_desc).value if c_desc else None
            if not norm(desc):
                continue
            tid = ws.cell(r, c_test_no).value
            if tid is not None and not isinstance(tid, (int, float)):
                if str(tid).strip():
                    continue

            mode_val = ws.cell(r, c_mode).value if c_mode else ""
            wave_val = ws.cell(r, c_wave).value if c_wave else ""
            if norm(mode_val):
                last_mode = mode_val
            else:
                mode_val = last_mode
            if norm(wave_val):
                last_wave = wave_val
            else:
                wave_val = last_wave

            res = infer_spec_update_step(
                description=desc,
                test_condition=ws.cell(r, c_cond).value if c_cond else "",
                mode=mode_val,
                waveform=wave_val,
                freq_col=ws.cell(r, c_freq).value if c_freq else "",
                pin_col=ws.cell(r, c_pin).value if c_pin else "",
                pout_col=ws.cell(r, c_pout).value if c_pout else "",
            )
            records.append(
                {
                    "row": r,
                    "test_id": tid,
                    "desc": desc,
                    "group_key": list(res["group_key"]),
                    "metric": res["metric"],
                    "suggested_mapping": res["step_name"],
                }
            )

print(
    json.dumps(
        {
            "input": args.input,
            "sheet": ws.title,
            "detected_style": style,
            "record_count": len(records),
            "records": records[:5000],
        },
        ensure_ascii=True,
        indent=2,
    )
)
