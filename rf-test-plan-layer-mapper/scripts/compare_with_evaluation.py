import argparse, json
from pathlib import Path
from openpyxl import load_workbook
from mapping_rules import detect_style_from_sheet


def find_text_in_row(ws, row, target):
    t = target.strip().lower()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row, c).value
        if isinstance(v, str) and v.strip().lower() == t:
            return c
    return None


def find_any_text_in_row(ws, row, targets):
    for target in targets:
        col = find_text_in_row(ws, row, target)
        if col is not None:
            return col
    return None


def mam(ws, col):
    m = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= col <= mr.max_col:
            for rr in range(mr.min_row, mr.max_row + 1):
                m[rr] = mr.min_row
    return m


ap = argparse.ArgumentParser()
ap.add_argument("--original", required=True)
ap.add_argument("--evaluation", required=True)
ap.add_argument("--sheet")
ap.add_argument("--output", required=True)
ap.add_argument("--delta-json", required=True)
args = ap.parse_args()
wbo = load_workbook(args.original)
wbe = load_workbook(args.evaluation)
ws_o = wbo[args.sheet] if args.sheet else wbo[wbo.sheetnames[0]]
ws_e = (
    wbe[args.sheet]
    if args.sheet and args.sheet in wbe.sheetnames
    else wbe[wbe.sheetnames[0]]
)
style = detect_style_from_sheet(ws_o, wbo.sheetnames, ws_o.title)
changes = []
if style == "E6":
    ni_o = find_any_text_in_row(ws_o, 2, ("NI Evalution", "NI Evaluation"))
    if ni_o is None:
        idx = (find_text_in_row(ws_o, 2, "Test Name") or 2) + 1
        ws_o.insert_cols(idx, amount=3)
        ws_o.cell(2, idx).value = "NI Evaluation"
        ws_o.cell(2, idx + 1).value = "TestTime_1site"
        ws_o.cell(2, idx + 2).value = "TestTime_4Site"
        ni_o = idx
    ni_e = find_any_text_in_row(ws_e, 2, ("NI Evalution", "NI Evaluation")) or 3
    for c_off in range(3):
        co = ni_o + c_off
        ce = ni_e + c_off
        # clear merges in target col
        for mr in list(ws_o.merged_cells.ranges):
            if mr.min_col <= co <= mr.max_col:
                try:
                    ws_o.unmerge_cells(str(mr))
                except:
                    pass
        for r in range(1, ws_e.max_row + 1):
            ws_o.cell(r, co).value = ws_e.cell(r, ce).value
        for mr in ws_e.merged_cells.ranges:
            if mr.min_col <= ce <= mr.max_col:
                ws_o.merge_cells(
                    start_row=mr.min_row,
                    start_column=co,
                    end_row=mr.max_row,
                    end_column=co,
                )
    for r in range(6, ws_e.max_row + 1):
        v = ws_e.cell(r, ni_e).value
        if v not in [None, ""]:
            changes.append({"row": r, "target_row": r, "column": ni_o, "new": str(v)})
    wbo.save(args.output)
    Path(args.delta_json).write_text(
        json.dumps(
            {"style": style, "sheet": ws_o.title, "changes": changes},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "output": args.output,
                "delta_json": args.delta_json,
                "style": style,
                "sheet": ws_o.title,
                "change_count": len(changes),
            },
            ensure_ascii=False,
        )
    )
    raise SystemExit(0)
if style == "E5":
    map_o = (
        find_text_in_row(ws_o, 1, "test step")
        or find_text_in_row(ws_o, 1, "Test Step")
        or 3
    )
    map_e = (
        find_text_in_row(ws_e, 1, "test step")
        or find_text_in_row(ws_e, 1, "Test Step")
        or 3
    )
    ao, ae = mam(ws_o, map_o), mam(ws_e, map_e)
    seen = set()
    max_row = max(ws_o.max_row, ws_e.max_row)
    for r in range(1, max_row + 1):
        re = ae.get(r, r)
        val = ws_e.cell(re, map_e).value if re <= ws_e.max_row else None
        if val not in [None, ""]:
            ro = ao.get(r, r)
            if (ro, map_o) not in seen:
                ws_o.cell(ro, map_o).value = val
                seen.add((ro, map_o))
                changes.append(
                    {"row": r, "target_row": ro, "column": map_o, "new": str(val)}
                )
    wbo.save(args.output)
    Path(args.delta_json).write_text(
        json.dumps(
            {"style": style, "sheet": ws_o.title, "changes": changes},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "output": args.output,
                "delta_json": args.delta_json,
                "style": style,
                "sheet": ws_o.title,
                "change_count": len(changes),
            },
            ensure_ascii=False,
        )
    )
    raise SystemExit(0)
raise SystemExit(f"Unsupported compare style: {style}")
