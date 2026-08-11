import argparse, json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from mapping_rules import (
    detect_style_from_sheet,
    find_semantic_columns,
    infer_e5_from_block,
    infer_e6_family,
    infer_e6_times,
    infer_limit_rf_mapping,
    infer_spec_update_step,
    norm,
)

Y = PatternFill("solid", fgColor="FFF2CC")
G = PatternFill("solid", fgColor="D9EAD3")


def find_text_in_row(ws, row, target):
    t = target.strip().lower()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row, c).value
        if isinstance(v, str) and v.strip().lower() == t:
            return c
    return None


def find_any_text_in_row(ws, row, targets):
    for target in targets:
        col = find_text_in_row(ws, row, target)
        if col is not None:
            return col
    return None


def merge_contiguous(ws, cols, start_row):
    main = cols[0]
    runs = []
    start = None
    cur = ""
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, main).value
        v = "" if v is None else str(v).strip()
        if start is None:
            start = r
            cur = v
        elif v != cur:
            if cur:
                runs.append((start, r - 1, cur))
            start = r
            cur = v
    if cur:
        runs.append((start, ws.max_row, cur))
    for s, e, v in runs:
        for c in cols:
            if e > s:
                ws.merge_cells(start_row=s, start_column=c, end_row=e, end_column=c)
            ws.cell(s, c).alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="left"
            )
    return runs


def mam(ws, col):
    m = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= col <= mr.max_col:
            for rr in range(mr.min_row, mr.max_row + 1):
                m[rr] = mr.min_row
    return m


def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


ap = argparse.ArgumentParser()
ap.add_argument("--input", required=True)
ap.add_argument("--sheet")
ap.add_argument("--output", required=True)
args = ap.parse_args()
wb = load_workbook(args.input)
ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
style = detect_style_from_sheet(ws, wb.sheetnames, ws.title)

if style == "LIMIT_RF":
    header_row = None
    for r in range(1, 10):
        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, 18)]
        if "test id" in vals and "Item" in vals and "Test Description" in vals:
            header_row = r
            break
    if header_row is None:
        raise SystemExit("LIMIT_RF header row not found")
    insert_col = 3
    ws.insert_cols(insert_col)
    ws.cell(header_row, insert_col).value = "Predicted Test Step"
    ws.cell(header_row, insert_col).fill = G
    ws.cell(header_row, insert_col).font = Font(bold=True)
    ws.cell(header_row, insert_col).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    sub = header_row + 1
    ws.cell(sub, insert_col).value = "skill v5.4 draft mapping"
    ws.cell(sub, insert_col).fill = G
    ws.cell(sub, insert_col).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    def get(r, c):
        return ws.cell(r, c if c < insert_col else c + 1).value

    start = sub + 1
    current_section = ""
    prev = None
    for r in range(start, ws.max_row + 1):
        tid = get(r, 1)
        item = get(r, 2)
        desc = get(r, 3)
        if not is_num(tid):
            section = " ".join(
                [norm(get(r, 1)), norm(get(r, 2)), norm(get(r, 3))]
            ).strip()
            if section:
                current_section = section
                prev = None
            continue
        band = get(r, 8)
        tx = get(r, 9)
        freq = get(r, 10)
        wave = get(r, 11)
        pwr = get(r, 12)
        m = infer_limit_rf_mapping(
            item, desc, band, tx, freq, wave, pwr, current_section, prev
        )
        if m:
            ws.cell(r, insert_col).value = m
            ws.cell(r, insert_col).fill = Y
            ws.cell(r, insert_col).alignment = Alignment(wrap_text=True, vertical="top")
            prev = m
    merge_contiguous(ws, (insert_col,), start)
    ws.column_dimensions["C"].width = 44
    ws.freeze_panes = f"D{start}"
    wb.save(args.output)
    print(args.output)
    raise SystemExit(0)

if style == "E6":
    ni_col = find_any_text_in_row(ws, 2, ("NI Evalution", "NI Evaluation"))
    if ni_col is None:
        idx = (find_text_in_row(ws, 2, "Test Name") or 2) + 1
        ws.insert_cols(idx, amount=3)
        for off, h in enumerate(["NI Evaluation", "TestTime_1site", "TestTime_4Site"]):
            ws.cell(2, idx + off).value = h
            ws.cell(2, idx + off).fill = G
            ws.cell(2, idx + off).font = Font(bold=True)
        ws.cell(3, idx).value = "Module Type_NI"
        ws.cell(3, idx + 1).value = "TestTime_1site"
        ws.cell(3, idx + 2).value = "TestTime_4Site"
        ni_col = idx
    t1_col = ni_col + 1
    t4_col = ni_col + 2

    def get(r, c):
        # if columns inserted, original cols >=3 shifted by +3
        return (
            ws.cell(r, c if c < ni_col else c + 3).value
            if ni_col == 3
            else ws.cell(r, c).value
        )

    # use workbook after insertion explicitly by column positions
    start_row = 6
    for r in range(start_row, ws.max_row + 1):
        n = ws.cell(r, 2).value
        if (
            isinstance(ws.cell(r, 1).value, (int, float))
            and isinstance(n, str)
            and n.strip()
        ):
            setting = ws.cell(r, 11).value if ni_col == 3 else ws.cell(r, 8).value
            pout = ws.cell(r, 13).value if ni_col == 3 else ws.cell(r, 10).value
            freq = ws.cell(r, 14).value if ni_col == 3 else ws.cell(r, 11).value
            wave = ws.cell(r, 15).value if ni_col == 3 else ws.cell(r, 12).value
            fam = infer_e6_family(
                n, str(setting or ""), pout, str(freq or ""), str(wave or "")
            )
            if fam:
                ws.cell(r, ni_col).value = fam
                ws.cell(r, ni_col).fill = Y
                t1, t4 = infer_e6_times(fam)
                if t1 is not None:
                    ws.cell(r, t1_col).value = t1
                    ws.cell(r, t1_col).fill = Y
                if t4 is not None:
                    ws.cell(r, t4_col).value = t4
                    ws.cell(r, t4_col).fill = Y
    merge_contiguous(ws, (ni_col, t1_col, t4_col), start_row)
    wb.save(args.output)
    print(args.output)
    raise SystemExit(0)

if style == "E5":
    map_col = find_text_in_row(ws, 1, "test step") or find_text_in_row(
        ws, 1, "Test Step"
    )
    if map_col is None:
        idx = (find_text_in_row(ws, 1, "TestNameMod") or 2) + 1
        ws.insert_cols(idx)
        ws.cell(1, idx).value = "test step"
        ws.cell(1, idx).fill = G
        ws.cell(1, idx).font = Font(bold=True)
        map_col = idx
    anchors = sorted(set(mam(ws, map_col).values()))
    if not anchors:
        anchors = []
        r = 1
        while r <= ws.max_row:
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            if not isinstance(a, (int, float)) and b is not None and r > 2:
                nr = r + 1
                while nr <= ws.max_row and not isinstance(
                    ws.cell(nr, 1).value, (int, float)
                ):
                    nr += 1
                if nr <= ws.max_row:
                    anchors.append(nr)
                r = nr
            else:
                r += 1
    for anchor in anchors:
        first = ws.cell(anchor, 2).value
        note = None
        for r in range(anchor, 1, -1):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            if not isinstance(a, (int, float)) and b is not None:
                note = b
                break
        step = infer_e5_from_block(first, note)
        if step:
            ws.cell(anchor, map_col).value = step
            ws.cell(anchor, map_col).fill = Y
            ws.cell(anchor, map_col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    wb.save(args.output)
    print(args.output)
    raise SystemExit(0)

if style == "SPEC_UPDATE":
    sem = find_semantic_columns(ws)
    if sem is None:
        raise SystemExit("SPEC_UPDATE semantic columns not found")

    cols = sem["columns"]
    header_row = sem["header_row"]
    desc_col = cols.get("description")
    if desc_col is None:
        raise SystemExit("Description column not found")

    map_col = find_text_in_row(ws, header_row, "Predicted Test Step")
    if map_col is None:
        map_col = desc_col + 1
        ws.insert_cols(map_col)
        ws.cell(header_row, map_col).value = "Predicted Test Step"
        ws.cell(header_row, map_col).fill = G
        ws.cell(header_row, map_col).font = Font(bold=True)
        ws.cell(header_row, map_col).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    start_row = header_row + 1

    def shifted(col):
        if col is None:
            return None
        return col if col < map_col else col + 1

    c_cond = shifted(cols.get("condition"))
    c_mode = shifted(cols.get("mode"))
    c_wave = shifted(cols.get("waveform"))
    c_freq = shifted(cols.get("freq"))
    c_pin = shifted(cols.get("pin"))
    c_pout = shifted(cols.get("pout"))
    c_test_no = shifted(cols.get("test_no"))
    c_desc = shifted(desc_col)

    group_step = {}
    last_mode = ""
    last_wave = ""
    for r in range(start_row, ws.max_row + 1):
        desc = ws.cell(r, c_desc).value
        if not norm(desc):
            continue
        tno = ws.cell(r, c_test_no).value if c_test_no else ws.cell(r, 1).value
        if tno is not None and not is_num(tno) and str(tno).strip():
            continue

        mode_val = ws.cell(r, c_mode).value if c_mode else ""
        wave_val = ws.cell(r, c_wave).value if c_wave else ""
        mode_norm = norm(mode_val)
        wave_norm = norm(wave_val)
        if mode_norm:
            last_mode = mode_val
        else:
            mode_val = last_mode
        if wave_norm:
            last_wave = wave_val
        else:
            wave_val = last_wave

        res = infer_spec_update_step(
            description=desc,
            test_condition=ws.cell(r, c_cond).value if c_cond else "",
            mode=mode_val,
            waveform=wave_val,
            freq_col=ws.cell(r, c_freq).value if c_freq else "",
            pin_col=ws.cell(r, c_pin).value if c_pin else "",
            pout_col=ws.cell(r, c_pout).value if c_pout else "",
        )
        key = res["group_key"]
        step = group_step.get(key)
        if step is None:
            step = res["step_name"]
            group_step[key] = step

        ws.cell(r, map_col).value = step
        ws.cell(r, map_col).fill = Y
        ws.cell(r, map_col).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions[ws.cell(header_row, map_col).column_letter].width = 48
    merge_contiguous(ws, (map_col,), start_row)
    ws.freeze_panes = ws.cell(start_row, map_col + 1).coordinate
    wb.save(args.output)
    print(args.output)
    raise SystemExit(0)

raise SystemExit(f"Unsupported style: {style}")
