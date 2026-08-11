#!/usr/bin/env python3
"""Deterministic Test Plan to TSRU workbook and sequence generator."""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import subprocess
import sys
import xml.etree.ElementTree as ET
from collections import OrderedDict
from pathlib import Path
from typing import Dict, List, Tuple, Iterable, Any

from openpyxl import Workbook, load_workbook

PINMAP_NS = {"pm": "http://www.ni.com/TestStand/SemiconductorModule/PinMap.xsd"}
MEASUREMENT_SUFFIXES = [
    "_OutputVoltage",
    "_InputVoltage",
    "_Leakage",
    "_DNL",
    "_INL",
    "_OffsetError",
    "_GainError",
    "_InputResistance",
]
UNIT_PATTERNS = [
    r"\(V\)",
    r"\(uA\)",
    r"\(LSB\)",
    r"\(mA\)",
    r"\(mV\)",
    r"\(KHz\)",
    r"\(kHz\)",
    r"\(khz\)",
    r"\(ohm\)",
    r"\s+MAX$",
    r"\s+MIN$",
    r"_MAX$",
    r"_MIN$",
]
FIELD_MAP = {
    "Test Name": ["TestName", "Test Name", "Test"],
    "Pin": ["Pin"],
    "L_Limit": ["Low Limit", "LowLimit", "Min", "Low"],
    "Low Limit": ["Low Limit", "LowLimit", "Min", "Low"],
    "H_Limit": ["High Limit", "HighLimit", "Max", "High"],
    "High Limit": ["High Limit", "HighLimit", "Max", "High"],
    "Unit": ["Unit", "Units"],
    "Units": ["Units", "Unit"],
    "Test Number": ["TestNumber", "Test Number"],
}


def normalize(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def normalize_no_digits(value: Any) -> str:
    return re.sub(r"[^a-z]", "", str(value or "").lower())


def strip_units(text: str) -> str:
    value = str(text or "")
    for pattern in UNIT_PATTERNS:
        value = re.sub(pattern, "", value, flags=re.IGNORECASE)
    return value.strip().strip("_")


def parse_pin_groups(pinmap_path: Path) -> Dict[str, List[str]]:
    root = ET.parse(pinmap_path).getroot()
    groups: Dict[str, List[str]] = OrderedDict()
    for pin_group in root.findall(".//pm:PinGroup", PINMAP_NS):
        name = pin_group.get("name", "")
        pins = [pin_ref.get("pin", "") for pin_ref in pin_group.findall("pm:PinReference", PINMAP_NS)]
        groups[name] = [pin for pin in pins if pin]
    return groups


def read_user_plan(user_plan_path: Path) -> List[Dict[str, Any]]:
    if user_plan_path.suffix.lower() == ".csv":
        with user_plan_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            return [normalize_record(row) for row in reader if any(v not in (None, "") for v in row.values())]

    workbook = load_workbook(user_plan_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if not row or not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        record = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        records.append(normalize_record(record))
    return records


def normalize_record(record: Dict[str, Any]) -> Dict[str, Any]:
    normalized = dict(record)
    for key in list(normalized.keys()):
        value = normalized[key]
        if isinstance(value, str):
            normalized[key] = value.strip()
    return normalized


def read_template(template_path: Path) -> Tuple[Dict[str, List[Dict[str, Any]]], Dict[str, List[str]], Dict[str, List[Dict[str, Any]]]]:
    workbook = load_workbook(template_path, data_only=True)
    template_rows: Dict[str, List[Dict[str, Any]]] = OrderedDict()
    template_headers: Dict[str, List[str]] = OrderedDict()
    template_defaults: Dict[str, List[Dict[str, Any]]] = OrderedDict()

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            template_headers[sheet_name] = []
            template_rows[sheet_name] = []
            template_defaults[sheet_name] = []
            continue

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        defaults: List[Dict[str, Any]] = []
        for row in rows[1:]:
            if not row or not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue
            defaults.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})

        template_headers[sheet_name] = headers
        template_rows[sheet_name] = []
        template_defaults[sheet_name] = defaults
    return template_rows, template_headers, template_defaults


def is_leakage_or_continuity(record: Dict[str, Any]) -> bool:
    name = str(record.get("TestName", "") or record.get("SequenceName", "")).lower()
    return "leakage" in name or "continuity" in name


def same_pin_group(pin_a: str, pin_b: str, pin_groups: Dict[str, List[str]]) -> bool:
    for pins in pin_groups.values():
        if pin_a in pins and pin_b in pins:
            return True
    return False


def get_grouping_key(record: Dict[str, Any]) -> str:
    for key in ["StepName", "SequenceName", "TestName"]:
        value = record.get(key)
        if value:
            return str(value).strip()
    return ""


def group_test_cases(records: List[Dict[str, Any]], pin_groups: Dict[str, List[str]]) -> List[List[Dict[str, Any]]]:
    chunks: List[List[Dict[str, Any]]] = []
    current_chunk: List[Dict[str, Any]] = []

    for index, record in enumerate(records):
        current_chunk.append(record)

        if len(current_chunk) < 5:
            continue

        if index + 1 < len(records):
            next_record = records[index + 1]
            key = get_grouping_key(record)
            next_key = get_grouping_key(next_record)

            if key == next_key and key != "":
                if is_leakage_or_continuity(record):
                    if same_pin_group(str(record.get("Pin", "")), str(next_record.get("Pin", "")), pin_groups):
                        continue
                if len(current_chunk) < 10:
                    continue

        chunks.append(current_chunk)
        current_chunk = []

    if current_chunk:
        chunks.append(current_chunk)
    return chunks


def expand_test_cases(chunk: List[Dict[str, Any]], pin_groups: Dict[str, List[str]]) -> List[Dict[str, Any]]:
    expanded: List[Dict[str, Any]] = []
    for record in chunk:
        pin = str(record.get("Pin", "") or "")
        sequence = str(record.get("SequenceName", "") or "")
        is_group = pin in pin_groups
        is_not_adc = "ADC" not in sequence.upper()

        if is_group and is_not_adc:
            group_pins = pin_groups[pin]
            base_number = safe_int(record.get("TestNumber"), 0)
            for offset, individual_pin in enumerate(group_pins):
                new_record = dict(record)
                test_name = str(record.get("TestName", "") or "")
                new_record["TestName"] = f"{test_name}_{individual_pin}" if test_name else individual_pin
                new_record["Pin"] = individual_pin
                new_record["TestNumber"] = base_number + offset
                expanded.append(new_record)
        else:
            expanded.append(dict(record))
    return expanded


def enrich_test_name_base(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    for record in records:
        test_name = str(record.get("TestName", "") or "")
        pin = str(record.get("Pin", "") or "")
        base = test_name

        if pin and base.endswith(f"_{pin}"):
            base = base[: -len(f"_{pin}")]

        for suffix in MEASUREMENT_SUFFIXES:
            if base.endswith(suffix):
                base = base[: -len(suffix)]
                break

        base = strip_units(base)
        base = base.replace(" ", "_").strip("_")
        if not base:
            base = test_name.replace(" ", "_").strip("_")
        record["Test Name Base"] = base
    return records


def enrich_pin_groups(records: List[Dict[str, Any]], pin_groups: Dict[str, List[str]]) -> List[Dict[str, Any]]:
    pin_to_groups: Dict[str, List[Tuple[str, int]]] = {}
    for group_name, pins in pin_groups.items():
        for pin in pins:
            pin_to_groups.setdefault(pin, []).append((group_name, len(pins)))

    for record in records:
        pin = str(record.get("Pin", "") or "").strip()
        if not pin:
            record["Pin Groups"] = ""
            continue
        if pin in pin_to_groups:
            candidates = sorted(pin_to_groups[pin], key=lambda item: item[1], reverse=True)
            record["Pin Groups"] = candidates[0][0]
        else:
            record["Pin Groups"] = pin
    return records


def select_worksheet(record: Dict[str, Any], template_headers: Dict[str, List[str]]) -> str:
    sequence = str(record.get("SequenceName", record.get("StepName", "")) or "")
    normalized_sequence = normalize(sequence)
    best_match = None
    best_score = -1
    for sheet_name in template_headers:
        normalized_sheet = normalize(sheet_name)
        if normalized_sequence in normalized_sheet or normalized_sheet in normalized_sequence:
            score = len(set(normalized_sequence) & set(normalized_sheet))
            if score > best_score:
                best_score = score
                best_match = sheet_name
    return best_match or next(iter(template_headers))


def select_template_defaults(record: Dict[str, Any], sheet_name: str, template_defaults: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not template_defaults:
        return {}
    if len(template_defaults) == 1:
        return template_defaults[0]

    record_sequence = str(record.get("SequenceName", "") or "")
    record_step = str(record.get("StepName", "") or "")
    record_test = str(record.get("TestName", "") or "")
    record_base = str(record.get("Test Name Base", "") or "")
    record_publish = strip_units(record_test)
    record_pin_group = str(record.get("Pin Groups", "") or "")

    best_row = template_defaults[0]
    best_score = -1

    for row in template_defaults:
        row_publish = str(row.get("Published Data ID", "") or "")
        row_source = str(row.get("Source Step", "") or "")
        row_pin_groups = str(row.get("pinsOrPinGroups", row.get("Pins and Pin Groups", "")) or "")
        row_dc = str(row.get("dcPowerPinNames", "") or "")
        row_pattern = str(row.get("patternPinNames", "") or "")

        score = 0
        if normalize(record_step) and normalize(record_step) == normalize(row_source):
            score += 100
        if normalize_no_digits(record_step) and normalize_no_digits(record_step) == normalize_no_digits(row_source):
            score += 80
        if normalize(record_publish) and normalize(record_publish) == normalize(row_publish):
            score += 70
        if normalize(record_publish) and normalize(record_publish) in normalize(row_publish):
            score += 40
        if normalize(row_publish) and normalize(row_publish) in normalize(record_publish):
            score += 35
        if "forcevoltage" in normalize(record_step) and "forcevoltage" in normalize(row_source):
            score += 120
        if "leakage" in normalize(record_test + record_sequence) and "leakage" in normalize(row_publish + row_source):
            score += 60
        if "continuity" in normalize(record_test + record_sequence) and "continuity" in normalize(row_publish + row_source):
            score += 60
        if "offset" in normalize(record_test) and ("oe" in normalize(row_publish) or "offset" in normalize(row_publish)):
            score += 30
        if "gain" in normalize(record_test) and ("ge" in normalize(row_publish) or "gain" in normalize(row_publish)):
            score += 30
        if "dnl" in normalize(record_test) and "dnl" in normalize(row_publish):
            score += 30
        if "inl" in normalize(record_test) and "inl" in normalize(row_publish):
            score += 30
        if record_pin_group and normalize(record_pin_group) in normalize(row_dc + row_pattern + row_pin_groups):
            score += 20
        if sheet_name.lower().startswith("adc") and row_source:
            score += 5

        if score > best_score:
            best_score = score
            best_row = row

    return best_row


def build_row(record: Dict[str, Any], target_headers: List[str], template_default: Dict[str, Any]) -> Dict[str, Any]:
    row: Dict[str, Any] = OrderedDict()
    for header in target_headers:
        value = ""
        template_value = template_default.get(header, "") if template_default else ""
        if is_non_empty(template_value):
            value = template_value
        elif header in record and is_non_empty(record[header]):
            value = record[header]
        else:
            for key in FIELD_MAP.get(header, []):
                if key in record and is_non_empty(record[key]):
                    value = record[key]
                    break
        row[header] = value
    apply_special_fields(row, record, target_headers, template_default)
    return row


def apply_special_fields(row: Dict[str, Any], record: Dict[str, Any], target_headers: List[str], template_default: Dict[str, Any]) -> None:
    test_name = str(record.get("TestName", "") or "")
    test_name_base = str(record.get("Test Name Base", "") or "")
    pin_groups = str(record.get("Pin Groups", "") or "")
    source_step = str(row.get("Source Step", "") or record.get("StepName", "") or "").strip()
    pin = str(record.get("Pin", "") or "")
    sequence = str(record.get("SequenceName", "") or "")

    if "Published Data ID" in target_headers and not is_non_empty(row.get("Published Data ID")):
        row["Published Data ID"] = strip_units(test_name)

    if "Source Step" in target_headers and not is_non_empty(row.get("Source Step")) and record.get("StepName"):
        row["Source Step"] = record.get("StepName")
        source_step = str(row.get("Source Step", "") or "")

    if "Pins and Pin Groups" in target_headers:
        row["Pins and Pin Groups"] = format_pin_group(pin_groups)

    if "pinsOrPinGroups" in target_headers:
        if not is_non_empty(row.get("pinsOrPinGroups")) and pin_groups:
            row["pinsOrPinGroups"] = format_pin_group(pin_groups)

    if "Pin" in target_headers and pin:
        row["Pin"] = pin

    if "voltageLevel" in target_headers and str(record.get("StepName", "") or "").strip().lower() == "force voltage":
        row["voltageLevel"] = record.get("Value")

    if "adcChannel" in target_headers and not is_non_empty(row.get("adcChannel")) and pin:
        match = re.search(r"(\d+)", pin)
        if match:
            row["adcChannel"] = int(match.group(1))

    if "Output Step" in target_headers:
        if str(record.get("StepName", "") or "").strip().lower() == "force voltage":
            voltage_value = record.get("Value")
            suffix = str(abs(int(voltage_value))) if voltage_value is not None and str(voltage_value) != "" else ""
            row["Output Step"] = f"Force_Voltage_Test{suffix}" if suffix else "Force_Voltage"
        elif source_step:
            clean_group = clean_pin_group(pin_groups)
            if clean_group and test_name_base:
                row["Output Step"] = f"{source_step}_{test_name_base}_{clean_group}"
            elif test_name_base:
                row["Output Step"] = f"{source_step}_{test_name_base}"

    if "Test Name" in target_headers and test_name:
        row["Test Name"] = test_name

    if "Low Limit" in target_headers and not is_non_empty(row.get("Low Limit")) and is_non_empty(record.get("LowLimit")):
        row["Low Limit"] = record.get("LowLimit")
    if "High Limit" in target_headers and not is_non_empty(row.get("High Limit")) and is_non_empty(record.get("HighLimit")):
        row["High Limit"] = record.get("HighLimit")
    if "Units" in target_headers and not is_non_empty(row.get("Units")) and is_non_empty(record.get("Units")):
        row["Units"] = record.get("Units")

    # Preserve template value only rule when already set.
    if "pinsOrPinGroups" in target_headers and is_non_empty(template_default.get("pinsOrPinGroups")):
        row["pinsOrPinGroups"] = template_default.get("pinsOrPinGroups")


def post_process(merged_results: Dict[str, List[Dict[str, Any]]], enriched_chunks: List[List[Dict[str, Any]]]) -> None:
    test_name_base_lookup: Dict[str, str] = {}
    for chunk in enriched_chunks:
        for record in chunk:
            test_name = str(record.get("TestName", "") or "")
            test_name_base = str(record.get("Test Name Base", "") or "")
            if test_name and test_name_base:
                test_name_base_lookup[test_name] = test_name_base

    for _, rows in merged_results.items():
        variable_counters: Dict[str, int] = {}
        for row in rows:
            if "Pins and Pin Groups" in row:
                row["Pins and Pin Groups"] = format_pin_group(row.get("Pins and Pin Groups", ""))

            if "Output Step" in row:
                base = test_name_base_lookup.get(str(row.get("Test Name", "") or ""), "")
                source_step = str(row.get("Source Step", "") or "").strip()
                group = clean_pin_group(row.get("Pins and Pin Groups", row.get("pinsOrPinGroups", "")))
                if source_step and base:
                    row["Output Step"] = f"{source_step}_{base}_{group}" if group else f"{source_step}_{base}"

            for key, value in list(row.items()):
                if isinstance(value, str) and "[index]" in value:
                    match = re.search(r"(.*?)\[index\]", value)
                    if match:
                        prefix = match.group(1)
                        variable_counters.setdefault(prefix, 0)
                        row[key] = value.replace("[index]", f"[{variable_counters[prefix]}]", 1)
                        variable_counters[prefix] += 1


def map_to_tsru(enriched_chunks: List[List[Dict[str, Any]]], template_headers: Dict[str, List[str]], template_defaults: Dict[str, List[Dict[str, Any]]]) -> Dict[str, List[Dict[str, Any]]]:
    merged: Dict[str, List[Dict[str, Any]]] = OrderedDict((sheet_name, []) for sheet_name in template_headers)
    for chunk in enriched_chunks:
        for record in chunk:
            sheet_name = select_worksheet(record, template_headers)
            defaults = select_template_defaults(record, sheet_name, template_defaults.get(sheet_name, []))
            merged[sheet_name].append(build_row(record, template_headers[sheet_name], defaults))
    return merged


def save_output(template_path: Path, merged_results: Dict[str, List[Dict[str, Any]]], template_headers: Dict[str, List[str]], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / "TSRU_Generated_TestPlan.xlsx"
    workbook = load_workbook(template_path)

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        headers = template_headers[sheet_name]
        for row_dict in merged_results.get(sheet_name, []):
            ws.append([row_dict.get(header, "") for header in headers])

    workbook.save(output_file)
    return output_file


def deploy_and_generate(generated_workbook: Path, tsru_modules_dir: Path, output_dir: Path) -> Tuple[Path, List[Path]]:
    deployed_workbook = tsru_modules_dir / "TestSpreadsheet.xlsx"
    shutil.copy2(generated_workbook, deployed_workbook)

    cmd_path = tsru_modules_dir / "Generate Seq.cmd"
    try:
        subprocess.run([str(cmd_path)], cwd=tsru_modules_dir, check=True, shell=True)
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(
            "TSRU workbook generation succeeded, but sequence generation failed while running "
            f"'{cmd_path}'. This commonly indicates a missing TSRU/TestStand license. "
            f"The generated workbook was preserved at '{generated_workbook}' and deployed to '{deployed_workbook}'."
        ) from exc

    generated_seq_files = sorted(tsru_modules_dir.glob("*.seq"))
    copied_outputs: List[Path] = []
    for seq_file in generated_seq_files:
        destination = output_dir / seq_file.name
        shutil.copy2(seq_file, destination)
        copied_outputs.append(destination)
    return deployed_workbook, copied_outputs


def format_pin_group(value: Any) -> str:
    clean = clean_pin_group(value)
    return f'{{"{clean}"}}' if clean else ""


def clean_pin_group(value: Any) -> str:
    return str(value or "").strip().strip("{}\"' ")


def is_non_empty(value: Any) -> bool:
    return value is not None and str(value).strip() not in {"", "nan", "None"}


def safe_int(value: Any, default: int) -> int:
    try:
        if value is None or value == "":
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def validate_file(path: Path, description: str) -> None:
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Missing required {description}: {path}")


def validate_tsru_modules(tsru_modules_dir: Path) -> None:
    if not tsru_modules_dir.exists() or not tsru_modules_dir.is_dir():
        raise FileNotFoundError(f"Missing required TSRU Modules folder: {tsru_modules_dir}")
    generate_seq = tsru_modules_dir / "Generate Seq.cmd"
    if not generate_seq.exists():
        raise FileNotFoundError(f"Missing required Generate Seq.cmd: {generate_seq}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert a test plan workbook into TSRU workbook and sequence files.")
    parser.add_argument("--template", required=True, help="Path to the TSRU template workbook")
    parser.add_argument("--plan", required=True, help="Path to the user test plan workbook or CSV")
    parser.add_argument("--pinmap", required=True, help="Path to the PinMap XML file")
    parser.add_argument("--tsru-modules", required=True, help="Path to the TSRU Modules folder")
    parser.add_argument("--output-dir", required=True, help="Path to the output folder")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    template_path = Path(args.template).resolve()
    plan_path = Path(args.plan).resolve()
    pinmap_path = Path(args.pinmap).resolve()
    tsru_modules_dir = Path(args.tsru_modules).resolve()
    output_dir = Path(args.output_dir).resolve()

    validate_file(template_path, "template workbook")
    validate_file(plan_path, "test plan file")
    validate_file(pinmap_path, "PinMap XML")
    validate_tsru_modules(tsru_modules_dir)

    pin_groups = parse_pin_groups(pinmap_path)
    user_records = read_user_plan(plan_path)
    if not user_records:
        raise ValueError("The test plan contains no records.")

    _, template_headers, template_defaults = read_template(template_path)
    chunks = group_test_cases(user_records, pin_groups)

    enriched_chunks: List[List[Dict[str, Any]]] = []
    for chunk in chunks:
        expanded = expand_test_cases(chunk, pin_groups)
        with_names = enrich_test_name_base(expanded)
        with_groups = enrich_pin_groups(with_names, pin_groups)
        enriched_chunks.append(with_groups)

    merged_results = map_to_tsru(enriched_chunks, template_headers, template_defaults)
    post_process(merged_results, enriched_chunks)
    generated_workbook = save_output(template_path, merged_results, template_headers, output_dir)
    deployed_workbook, copied_seq_files = deploy_and_generate(generated_workbook, tsru_modules_dir, output_dir)

    print(f"Generated workbook: {generated_workbook}")
    print(f"Deployed workbook: {deployed_workbook}")
    for seq_file in copied_seq_files:
        print(f"Generated sequence: {seq_file}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover
        print(f"Error: {exc}", file=sys.stderr)
        raise
